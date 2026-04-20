# app/routes/main.py
from flask import Blueprint, render_template, jsonify, request, url_for, redirect, send_file  # added redirect, send_file
from app import db
from app.models.tran import (
    Agency, FunctionalArea, Component, Vendor, IntegrationPoint,
    UpdateLog, Function, Standard, Tag, TagGroup, UserRole,
    integration_standard, component_integration,
    Configuration, ConfigurationProduct, Product, ProductVersion
)
from app.forms.forms import AgencyForm, VendorForm, ComponentForm
from app.auth import login_required, admin_required, get_updated_by
from app.utils.errors import (
    api_ok,
    json_error_response, json_success_response,
    html_error_fragment, html_success_fragment,
    json_form_error_response, json_validation_error_response
)
# removed import of AFI utility helpers (create_afi_with_optional_children, component_supports_function, etc.)
from sqlalchemy import func, case, distinct
from sqlalchemy.orm import joinedload
from sqlalchemy.exc import IntegrityError
from datetime import datetime, timedelta
import io  # for Excel streaming


main = Blueprint("main", __name__)

@main.route("/")
def index():
    from flask import session
    if session.get('user'):
        return redirect(url_for('main.dashboard'))
    return render_template("index.html")


@main.route("/dashboard")
def dashboard():
    from flask import session
    user = session.get('user', {})
    return render_template("dashboard.html", user=user)

# --- Basic pages ---
@main.route("/functional-areas")
def functional_areas_page():
    return render_template('functional_areas.html')

@main.get('/functional-areas/export.xlsx')
def export_functional_areas_excel():
    """Download an Excel export of Functional Areas with their Functions.

    Columns: Functional Area, Function, Criticality, # Components, # Agencies
    Styling: title row, header styling, borders, auto-width, frozen header, autofilter,
    conditional fill by criticality. Import of openpyxl is lazy to avoid import errors
    when the package isn't installed during unrelated operations/tests.
    """
    try:
        # Optional filter parity with the page (currently only a search box)
        search = (request.args.get('search') or '').strip()

        q = (FunctionalArea.query
             .options(joinedload(FunctionalArea.functions))
             .order_by(FunctionalArea.name.asc()))
        if search:
            q = q.filter(FunctionalArea.name.ilike(f"%{search}%"))
        areas = q.all()

        # Lazy import for safety in environments without openpyxl preinstalled
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except Exception as ie:
            return json_error_response(f"Excel export unavailable (dependency): {ie}", 500)

        wb = Workbook()
        ws = wb.active
        ws.title = "Functional Areas"

        # Styles
        title_font = Font(size=14, bold=True)
        subtitle_font = Font(size=10, color="6B7280")
        header_font = Font(color="FFFFFF", bold=True)
        header_fill = PatternFill("solid", fgColor="1F2937")  # slate-800
        thin = Side(style="thin", color="374151")  # slate-700
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Title and metadata
        ws["A1"] = "Functional Areas Export"
        ws["A1"].font = title_font
        ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
        ws["A2"].font = subtitle_font
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

        # Header row
        headers = [
            "Functional Area",
            "Function",
            "Criticality",
            "# Components",
            "# Agencies",
        ]
        ws.append([None])  # row 3 spacer so header lands at row 4 consistently
        ws.append(headers)  # this becomes row 4
        header_row_idx = 4
        for cell in ws[header_row_idx]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
            cell.border = border

        # Data rows
        row = header_row_idx + 1
        severity_order = {"high": 0, "medium": 1, "low": 2}
        crit_fills = {
            "high": PatternFill("solid", fgColor="B91C1C"),    # red-700
            "medium": PatternFill("solid", fgColor="B45309"),  # amber-700
            "low": PatternFill("solid", fgColor="065F46"),     # emerald-800
        }

        # Precompute agency counts per function using Configuration
        # Note: simple per-function query to keep it readable; acceptable for moderate sizes
        for area in areas:
            # Sort functions by criticality then name for consistency with print view
            functions = sorted(
                list(area.functions),
                key=lambda fx: (
                    severity_order.get(getattr(getattr(fx, 'criticality', None), 'value', 'medium'), 1),
                    (fx.name or '').lower(),
                ),
            )
            if not functions:
                # Emit an area line with em dash if no functions yet
                ws.cell(row=row, column=1, value=area.name)
                ws.cell(row=row, column=2, value="—")
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row, column=c).border = border
                row += 1
                continue

            for f in functions:
                # Component count via association
                try:
                    component_count = len(f.components)
                except Exception:
                    component_count = 0
                # Distinct agencies with configurations for this function
                agency_count = db.session.query(
                    func.count(func.distinct(Configuration.agency_id))
                ).filter(Configuration.function_id == f.id).scalar() or 0

                crit_val = getattr(getattr(f, 'criticality', None), 'value', None)
                crit_disp = crit_val.title() if crit_val else None

                ws.cell(row=row, column=1, value=area.name)
                ws.cell(row=row, column=2, value=f.name)
                crit_cell = ws.cell(row=row, column=3, value=crit_disp)
                ws.cell(row=row, column=4, value=component_count)
                ws.cell(row=row, column=5, value=agency_count)

                # Borders and criticality color badges (fill only the crit cell for subtlety)
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row, column=c).border = border
                    if c == 3 and crit_val in crit_fills:
                        ws.cell(row=row, column=c).fill = crit_fills[crit_val]
                        ws.cell(row=row, column=c).font = Font(color="FFFFFF", bold=True)
                        ws.cell(row=row, column=c).alignment = Alignment(horizontal="center")

                row += 1

        # Freeze header and add auto-filter
        ws.freeze_panes = f"A{header_row_idx + 1}"
        ws.auto_filter.ref = f"A{header_row_idx}:" + get_column_letter(len(headers)) + f"{row - 1}"

        # Auto-fit columns
        for col_idx in range(1, len(headers) + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for r in range(1, row):
                v = ws.cell(r, col_idx).value
                max_len = max(max_len, len(str(v)) if v is not None else 0)
            ws.column_dimensions[letter].width = min(max_len + 2, 48)

        # Stream workbook to response
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"functional_areas_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            buf,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        return json_error_response(f"Error generating export: {str(e)}", 500)

@main.get('/docs')
@login_required
def docs_index():
    import os, glob
    docs_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'docs')
    # Collect .md files (exclude hidden) and basic metadata
    files = []
    for path in sorted(glob.glob(os.path.join(docs_dir, '*.md'))):
        name = os.path.basename(path)
        try:
            with open(path, 'r', encoding='utf-8') as f:
                first = ''
                for _ in range(5):
                    line = f.readline()
                    if not line: break
                    if line.strip():
                        first = line.strip().lstrip('# ').strip()
                        break
                title = first or name
        except Exception:
            title = name
        files.append({'filename': name, 'title': title})
    # Default: show first doc
    selected = request.args.get('file') or (files[0]['filename'] if files else None)
    content_html = ''
    if selected:
        sel_path = os.path.join(docs_dir, selected)
        if os.path.isfile(sel_path):
            try:
                import markdown
                with open(sel_path, 'r', encoding='utf-8') as f:
                    raw = f.read()
                content_html = markdown.markdown(raw, extensions=['fenced_code', 'tables', 'toc'])
            except Exception as e:
                content_html = f"<p class='text-red-400 text-sm'>Error rendering: {e}</p>"
        else:
            content_html = "<p class='text-red-400 text-sm'>File not found.</p>"
    return render_template('docs.html', files=files, selected=selected, content_html=content_html)

# Print-friendly grid of all Functional Areas
@main.route("/functional-areas/print")
def functional_areas_print_page():
    try:
        areas = FunctionalArea.query.order_by(FunctionalArea.name.asc()).all()
        return render_template(
            'functional_areas_print.html',
            functional_areas=areas,
            title="Functional Areas – Print"
        )
    except Exception as e:
        # Fallback to empty list with error message embedded in page
        return render_template(
            'functional_areas_print.html',
            functional_areas=[],
            error=str(e),
            title="Functional Areas – Print"
        )

# Print-friendly page listing all functions organized by Functional Area
@main.route("/functions/print")
def functions_print_page():
    try:
        areas = (FunctionalArea.query
                 .options(joinedload(FunctionalArea.functions))
                 .order_by(FunctionalArea.name.asc())
                 .all())

        # decorate each function with quick counts and sort by criticality then name
        severity_order = {'high': 0, 'medium': 1, 'low': 2}
        for area in areas:
            for f in area.functions:
                try:
                    f.component_count = len(f.components)
                except Exception:
                    f.component_count = 0
                # distinct agencies using this function via configurations
                agency_count = db.session.query(func.count(func.distinct(Configuration.agency_id))) \
                    .filter(Configuration.function_id == f.id).scalar() or 0
                f.agency_count = agency_count
            # sorted list for display
            area.sorted_functions = sorted(
                list(area.functions),
                key=lambda fx: (
                    severity_order.get(getattr(getattr(fx, 'criticality', None), 'value', 'medium'), 1),
                    fx.name.lower()
                )
            )

        return render_template(
            'functions_print.html',
            functional_areas=areas,
            title="Functions – Print"
        )
    except Exception as e:
        return render_template(
            'functions_print.html',
            functional_areas=[],
            error=str(e),
            title="Functions – Print"
        )

@main.route('/reports')
def reports_page():
    # lightweight placeholder page
    return render_template('reports.html') if False else render_template('index.html')

@main.route("/components")
def components_page():
    """Components management page"""
    return render_template("components.html")

@main.route("/vendors")
def vendors_page():
    """Vendors management page"""
    return render_template("vendors.html")

# Health and utility endpoints
@main.route("/api/health")
def health_check():
    try:
        # Test database connection
        db.session.execute(db.text('SELECT 1'))
        return api_ok({
            "status": "ok",
            "timestamp": datetime.utcnow().isoformat(),
            "database": "connected"
        })
    except Exception as e:
        return json_error_response(f"Health check failed: {str(e)}", 500)

# ========== Functional Areas API (CRUD + fragments) ==========

@main.route('/api/functional-areas/list')
def functional_areas_list():
    try:
        search = (request.args.get('search') or '').strip()
        q = FunctionalArea.query
        if search:
            q = q.filter(FunctionalArea.name.ilike(f"%{search}%"))
        areas = q.order_by(FunctionalArea.name.asc()).all()
        return render_template('fragments/functional_area_list.html', functional_areas=areas)
    except Exception as e:
        return html_error_fragment(f"Error loading functional areas: {str(e)}")

@main.route('/api/functional-areas/<int:area_id>/details')
def functional_area_details(area_id):
    try:
        area = FunctionalArea.query.get_or_404(area_id)
        # decorate functions with counts used in template
        for f in area.functions:
            # component count from association
            try:
                f.component_count = len(f.components)
            except Exception:
                f.component_count = 0
            # agency count via configurations
            agency_count = db.session.query(func.count(func.distinct(Configuration.agency_id))) \
                .filter(Configuration.function_id == f.id).scalar() or 0
            f.agency_count = agency_count
        # sort functions by criticality severity then name
        severity_order = {'high': 0, 'medium': 1, 'low': 2}
        area.sorted_functions = sorted(
            list(area.functions),
            key=lambda fx: (severity_order.get(getattr(getattr(fx, 'criticality', None), 'value', 'medium'), 1), fx.name.lower())
        )
        return render_template('fragments/functional_area_details.html', functional_area=area)
    except Exception as e:
        return html_error_fragment(f"Error loading functional area details: {str(e)}")

@main.route('/api/functional-areas/form')
def functional_area_form():
    try:
        return render_template('fragments/functional_area_form.html', functional_area=None)
    except Exception as e:
        return html_error_fragment(f"Error loading form: {str(e)}")

@main.route('/api/functional-areas/<int:area_id>/form')
def functional_area_edit_form(area_id):
    try:
        area = FunctionalArea.query.get_or_404(area_id)
        return render_template('fragments/functional_area_form.html', functional_area=area)
    except Exception as e:
        return html_error_fragment(f"Error loading edit form: {str(e)}")

@main.route('/api/functional-areas', methods=['POST'])
@admin_required
def functional_area_create():
    try:
        name = (request.form.get('name') or '').strip()
        description = (request.form.get('description') or '').strip() or None
        if not name:
            return html_error_fragment('Name is required')
        area = FunctionalArea(name=name, description=description)
        db.session.add(area)
        db.session.commit()
        return html_success_fragment('Functional area created')
    except Exception as e:
        db.session.rollback()
        return html_error_fragment(f"Error creating functional area: {str(e)}")

@main.route('/api/functional-areas/<int:area_id>', methods=['PUT'])
@admin_required
def functional_area_update(area_id):
    try:
        area = FunctionalArea.query.get_or_404(area_id)
        name = (request.form.get('name') or '').strip()
        description = (request.form.get('description') or '').strip() or None
        if not name:
            return html_error_fragment('Name is required')
        area.name = name
        area.description = description
        db.session.commit()
        return html_success_fragment('Functional area updated')
    except Exception as e:
        db.session.rollback()
        return html_error_fragment(f"Error updating functional area: {str(e)}")

@main.route('/api/functional-areas/<int:area_id>', methods=['DELETE'])
@admin_required
def functional_area_delete(area_id):
    try:
        area = FunctionalArea.query.get_or_404(area_id)
        db.session.delete(area)
        db.session.commit()
        return html_success_fragment('Functional area deleted')
    except Exception as e:
        db.session.rollback()
        return html_error_fragment(f"Error deleting functional area: {str(e)}")

# Count endpoints for dashboard metrics
@main.route("/api/count/agencies")
def count_agencies():
    try:
        count = Agency.query.count()
        return str(count)
    except Exception as e:
        return "0"

@main.route("/api/count/functional-areas")
def count_functional_areas():
    try:
        count = FunctionalArea.query.count()
        return str(count)
    except Exception as e:
        return "0"

@main.route("/api/count/components")
def count_components():
    try:
        count = Component.query.count()
        return str(count)
    except Exception as e:
        return "0"

@main.route("/api/count/integration-points")
def count_integration_points():
    try:
        count = IntegrationPoint.query.count()
        return str(count)
    except Exception as e:
        return "0"

@main.route("/api/count/vendors")
def count_vendors():
    try:
        count = Vendor.query.count()
        return str(count)
    except Exception as e:
        return "0"

@main.route("/api/count/configurations")
def count_configurations():
    try:
        return str(Configuration.query.count())
    except Exception:
        return "0"

@main.route("/api/count/products")
def count_products():
    try:
        return str(Product.query.count())
    except Exception:
        return "0"


# Dashboard endpoints

@main.route("/api/dashboard/recent-configs")
def dashboard_recent_configs():
    """HTMX fragment: last 10 configurations for the dashboard activity feed."""
    try:
        configs = (
            Configuration.query
            .options(
                joinedload(Configuration.agency),
                joinedload(Configuration.function).joinedload(Function.functional_area),
                joinedload(Configuration.component),
                joinedload(Configuration.products).joinedload(ConfigurationProduct.product).joinedload(Product.vendor),
            )
            .order_by(Configuration.created_at.desc())
            .limit(10)
            .all()
        )
        return render_template('fragments/dashboard_recent_configs.html', configs=configs)
    except Exception as e:
        return html_error_fragment(f"Error loading recent configurations: {str(e)}")


@main.route("/api/dashboard/top-agencies")
def dashboard_top_agencies():
    """HTMX fragment: agencies ranked by configuration count."""
    try:
        rows = (
            db.session.query(Agency, func.count(Configuration.id).label('cfg_count'))
            .outerjoin(Configuration, Configuration.agency_id == Agency.id)
            .group_by(Agency.id)
            .order_by(func.count(Configuration.id).desc())
            .limit(8)
            .all()
        )
        return render_template('fragments/dashboard_top_agencies.html', rows=rows)
    except Exception as e:
        return html_error_fragment(f"Error loading top agencies: {str(e)}")


# Components endpoints
@main.route("/api/components/list")
def components_list():
    """Get all components with filtering (updated to use Configuration instead of AFI)."""
    try:
        functional_area = (request.args.get('functional_area') or '').strip()
        agency = (request.args.get('agency') or '').strip()
        status = (request.args.get('status') or '').strip()
        search = (request.args.get('search') or '').strip()
        # vendor filter removed (Component no longer tied to vendor)
        query = db.session.query(Component).distinct()
        if functional_area:
            query = (query
                     .join(Configuration, Configuration.component_id == Component.id)
                     .join(Function, Function.id == Configuration.function_id)
                     .join(FunctionalArea, FunctionalArea.id == Function.functional_area_id)
                     .filter(FunctionalArea.name == functional_area))
        if agency:
            query = (query
                     .join(Configuration, Configuration.component_id == Component.id)
                     .join(Agency, Agency.id == Configuration.agency_id)
                     .filter(Agency.name == agency))
        if status:
            query = (query
                     .join(Configuration, Configuration.component_id == Component.id)
                     .filter(Configuration.status == status))
        if search:
            name_like = f"%{search}%"
            query = query.filter(Component.name.ilike(name_like))
        query = query.order_by(Component.name.asc())
        components = query.all()
        view_components = []
        for component in components:
            agencies_using = (db.session.query(Agency.name)
                              .join(Configuration, Configuration.agency_id == Agency.id)
                              .filter(Configuration.component_id == component.id)
                              .distinct().limit(3).all())
            agencies_display = ", ".join([a.name for a in agencies_using]) or 'No agencies'
            if len(agencies_using) == 3:
                agencies_display += ' +more'
            functions_implemented = (db.session.query(Function.name)
                                     .join(Configuration, Configuration.function_id == Function.id)
                                     .filter(Configuration.component_id == component.id)
                                     .distinct().limit(3).all())
            functions_display = ", ".join([f.name for f in functions_implemented]) or 'No functions'
            if len(functions_implemented) == 3:
                functions_display += ' +more'
            # NEW: latest configuration version label & deployment date (for display on card)
            latest_cfg = (Configuration.query
                          .filter(Configuration.component_id == component.id)
                          .order_by(Configuration.deployment_date.desc().nullslast(), Configuration.updated_at.desc())
                          .first())
            version_label = getattr(latest_cfg, 'version_label', None) if latest_cfg else None
            deployment_date_str = ''
            if latest_cfg and latest_cfg.deployment_date:
                try:
                    deployment_date_str = latest_cfg.deployment_date.strftime('%Y-%m-%d')
                except Exception:
                    deployment_date_str = ''
            view_components.append(type('VC', (), {
                'id': component.id,
                'name': component.name,
                'is_composite': False,
                'status_indicator': 'green',
                'functions_display': functions_display,
                'vendor_name': '—',
                'agencies_display': agencies_display,
                'deployment_date_str': deployment_date_str,
                'version': version_label,  # keep legacy usage
                'version_label': version_label,
                'known_issues': None,
                'short_description': component.short_description,
            }))
        return render_template('fragments/component_list.html', components=view_components)
    except Exception as e:
        return html_error_fragment(f"Error loading components: {str(e)}")

@main.route("/api/components/<int:component_id>/details")
def component_details(component_id):
    """Updated component details using Configurations."""
    try:
        component = Component.query.get_or_404(component_id)
        configurations = (Configuration.query
                          .filter_by(component_id=component_id)
                          .join(Agency).join(Function).join(FunctionalArea)
                          .order_by(Agency.name, FunctionalArea.name, Function.name)
                          .all())
        agency_usage_html = ""
        if configurations:
            agency_usage_html = "<h4 class='font-medium text-white mb-3'>Agency Usage:</h4>"
            agencies = {}
            for c in configurations:
                agencies.setdefault(c.agency.name, []).append(c)
            for agency_name, cfgs in agencies.items():
                agency_usage_html += f'''<div class="mb-4"><h5 class="text-sm font-medium text-blue-400 mb-2">{agency_name}</h5><div class="space-y-2 ml-3">'''
                for cfg in cfgs:
                    status_color = "green" if cfg.status == "Active" else "yellow"
                    agency_usage_html += f'''<div class="flex items-center justify-between p-2 bg-slate-700/30 rounded"><div class="flex items-center space-x-2"><div class="w-2 h-2 bg-{status_color}-500 rounded-full"></div><span class="text-sm text-slate-300">{cfg.function.name}</span></div><div class="text-right"><span class="text-xs text-slate-500">{cfg.deployment_date.strftime('%Y-%m-%d') if cfg.deployment_date else 'No date'}</span>{f'<br><span class="text-xs text-slate-400">{cfg.version_label}</span>' if cfg.version_label else ''}</div></div>'''
                agency_usage_html += "</div></div>"
        else:
            agency_usage_html = "<p class='text-slate-400 text-sm'>No configuration usage tracked for this component.</p>"
        roles = ""
        if component.user_roles:
            roles = "<h4 class='font-medium text-white mb-2 mt-4'>User Roles:</h4><ul class='space-y-1'>" + \
                "".join([f'<li class="text-sm text-slate-300">• {r.role_name}: {r.description or "No description"}</li>' for r in component.user_roles]) + "</ul>"
        metadata = ""
        if component.additional_metadata:
            metadata = "<h4 class='font-medium text-white mb-2 mt-4'>Additional Information:</h4><ul class='space-y-1'>" + \
                "".join([f'<li class="text-sm text-slate-300">• {k.replace("_"," ").title()}: {v}</li>' for k,v in component.additional_metadata.items()]) + "</ul>"
        # NEW: include short_description and description
        short_desc_html = f"<p class='text-slate-300 mt-1'>{component.short_description}</p>" if getattr(component, 'short_description', None) else ""
        description_html = f"<div class='mt-4'><h3 class='font-medium text-white mb-2'>Description</h3><p class='text-slate-300 text-sm'>{component.description}</p></div>" if getattr(component, 'description', None) else ""
        html = f'''<div class="glass-effect rounded-xl p-6 border border-slate-700/50"><div class="flex items-center justify-between mb-4"><div class="flex items-center space-x-3"><h2 class="text-2xl font-bold text-white">{component.name}</h2></div><button class="px-3 py-1 bg-slate-700 hover:bg-slate-600 rounded text-sm transition-colors" onclick="clearComponentDetails()">✕ Close</button></div><div class="grid grid-cols-1 gap-6"><div><div class="space-y-2 text-sm">{short_desc_html}</div>{description_html}<div class="mt-6">{agency_usage_html}</div>{roles}{metadata}</div></div></div>'''
        return html
    except Exception as e:
        return html_error_fragment(f"Error loading component details: {str(e)}")

@main.route("/api/agencies/options")
def agencies_filter_options():
    """Get agency options for filter dropdowns (based on Configurations now)."""
    try:
        agencies = (db.session.query(Agency.name)
                    .join(Configuration, Configuration.agency_id == Agency.id)
                    .distinct()
                    .order_by(Agency.name)
                    .all())
        html = '<option value="">All Agencies</option>'
        for agency in agencies:
            html += f'<option value="{agency.name}">{agency.name}</option>'
        return html
    except Exception as e:
        return html_error_fragment(f"Error loading agency options: {str(e)}")

@main.route("/api/filter-options/functional-areas")
def functional_area_filter_options():
    """HTMX: functional area options (derived from active Configuration usage)."""
    try:
        fa_rows = (db.session.query(FunctionalArea.name)
                   .join(Function, Function.functional_area_id == FunctionalArea.id)
                   .join(Configuration, Configuration.function_id == Function.id)
                   .distinct()
                   .order_by(FunctionalArea.name)
                   .all())
        html = '<option value="">All Functional Areas</option>'
        for fa in fa_rows:
            html += f'<option value="{fa.name}">{fa.name}</option>'
        return html
    except Exception as e:
        return html_error_fragment(f"Error loading functional areas: {str(e)}")

@main.route('/api/vendors/filter-options/functional-areas')
def vendors_functional_area_filter_options():
    """Alias for vendors page expecting this endpoint."""
    return functional_area_filter_options()

@main.route('/api/vendors/filter-options/agencies')
def vendors_agency_filter_options():
    """Alias endpoint for vendor page to load agency options."""
    return agencies_filter_options()


@main.route('/api/integration/standards')
def integration_standards_list():
    """HTML fragment: list integration standards with simple usage counts (placeholder logic)."""
    try:
        # Basic list ordered by name; extend with real metrics later
        standards = Standard.query.order_by(Standard.name.asc()).all()
        if not standards:
            return '<p class="text-slate-500 text-sm">No standards defined.</p>'
        html_parts = []
        for s in standards:
            html_parts.append(f"<div class='flex items-center justify-between p-2 bg-slate-800/40 rounded text-sm'><span class='text-slate-300'>{s.name}</span></div>")
        return "\n".join(html_parts)
    except Exception as e:
        return html_error_fragment(f"Error loading standards: {str(e)}")

# -------- COMPONENT FORMS (vendor removed) ----------
@main.route("/api/components/form")
@login_required
def component_form():
    try:
        form = ComponentForm()
        return render_template('fragments/component_form.html', form=form, component=None)
    except Exception as e:
        return html_error_fragment(f"Error loading component form: {str(e)}")

@main.route("/api/components/<int:component_id>/form")
@login_required
def component_edit_form(component_id):
    try:
        component = Component.query.get_or_404(component_id)
        form = ComponentForm()
        form.populate_from_component(component)
        return render_template('fragments/component_form.html', form=form, component=component)
    except Exception as e:
        return html_error_fragment(f"Error loading component edit form: {str(e)}")

# NEW: Component create/update/delete endpoints
@main.route("/api/components", methods=['POST'])
@admin_required
def component_create():
    form = ComponentForm()
    if form.validate_on_submit():
        try:
            c = Component()
            form.populate_component(c)
            db.session.add(c)
            db.session.commit()
            return api_ok({"id": c.id, "message": f"Component '{c.name}' created"})
        except IntegrityError as ie:
            db.session.rollback()
            return json_error_response(f"Error creating component: {str(ie)}")
        except Exception as e:
            db.session.rollback()
            return json_error_response(f"Error creating component: {str(e)}")
    return json_form_error_response(form)

@main.route("/api/components/<int:component_id>", methods=['POST'])
@admin_required
def component_update(component_id):
    component = Component.query.get_or_404(component_id)
    form = ComponentForm()
    if form.validate_on_submit():
        try:
            form.populate_component(component)
            db.session.commit()
            return api_ok({"id": component.id, "message": f"Component '{component.name}' updated"})
        except Exception as e:
            db.session.rollback()
            return json_error_response(f"Error updating component: {str(e)}")
    return json_form_error_response(form)

@main.route("/api/components/<int:component_id>", methods=['DELETE'])
@admin_required
def component_delete(component_id):
    try:
        component = Component.query.get_or_404(component_id)
        # Prevent deletion if used in any configuration
        in_use = Configuration.query.filter_by(component_id=component_id).first() is not None
        if in_use:
            return json_error_response("Cannot delete component – it is referenced by one or more configurations.")
        db.session.delete(component)
        db.session.commit()
        return api_ok({"message": "Component deleted successfully"})
    except Exception as e:
        db.session.rollback()
        return json_error_response(f"Error deleting component: {str(e)}")

# -------- VENDORS LIST (repoint to Products & ConfigurationProduct) ----------
@main.route("/api/vendors/list")
def vendors_list():
    """Vendors list using product & configuration usage metrics."""
    try:
        search = request.args.get('search', '').strip()
        agency_filter = request.args.get('agency', '').strip()
        functional_area_filter = request.args.get('functional_area', '').strip()
        sort_by = request.args.get('sort', 'name')

        product_sub = db.session.query(
            Product.vendor_id.label('v_id'),
            func.count(Product.id).label('product_count')
        ).group_by(Product.vendor_id).subquery()

        usage_sub = db.session.query(
            Product.vendor_id.label('v_id'),
            func.count(func.distinct(Configuration.id)).label('usage_count')
        ).join(ConfigurationProduct, ConfigurationProduct.product_id == Product.id) \
         .join(Configuration, Configuration.id == ConfigurationProduct.configuration_id) \
         .group_by(Product.vendor_id).subquery()

        q = db.session.query(
            Vendor,
            func.coalesce(product_sub.c.product_count, 0).label('product_count'),
            func.coalesce(usage_sub.c.usage_count, 0).label('usage_count')
        ).outerjoin(product_sub, product_sub.c.v_id == Vendor.id) \
         .outerjoin(usage_sub, usage_sub.c.v_id == Vendor.id)

        if agency_filter or functional_area_filter:
            usage_filter_query = db.session.query(func.distinct(Vendor.id)) \
                .join(Product, Product.vendor_id == Vendor.id) \
                .join(ConfigurationProduct, ConfigurationProduct.product_id == Product.id) \
                .join(Configuration, Configuration.id == ConfigurationProduct.configuration_id)
            if agency_filter:
                usage_filter_query = usage_filter_query.join(Agency, Agency.id == Configuration.agency_id) \
                    .filter(Agency.name == agency_filter)
            if functional_area_filter:
                usage_filter_query = usage_filter_query.join(Function, Function.id == Configuration.function_id) \
                    .join(FunctionalArea, FunctionalArea.id == Function.functional_area_id) \
                    .filter(FunctionalArea.name == functional_area_filter)
            vendor_ids = [vid[0] for vid in usage_filter_query.all()]
            if vendor_ids:
                q = q.filter(Vendor.id.in_(vendor_ids))
            else:
                q = q.filter(db.text("1=0"))

        if search:
            q = q.filter(Vendor.name.ilike(f'%{search}%'))

        if sort_by in ('components', 'products'):
            q = q.order_by(func.coalesce(product_sub.c.product_count, 0).desc(), Vendor.name.asc())
        elif sort_by == 'recent':
            pv_sub = db.session.query(
                Product.vendor_id.label('v_id'),
                func.max(ProductVersion.release_date).label('latest_release')
            ).join(ProductVersion, ProductVersion.product_id == Product.id) \
             .group_by(Product.vendor_id).subquery()
            q = q.outerjoin(pv_sub, pv_sub.c.v_id == Vendor.id) \
                 .order_by(pv_sub.c.latest_release.desc().nullslast(), Vendor.name.asc())
        else:
            q = q.order_by(Vendor.name.asc())

        rows = q.all()
        for vendor, product_count, usage_count in rows:
            vendor.product_count = product_count
            vendor.usage_count = usage_count
        return render_template('fragments/vendor_list.html', vendors_with_counts=rows)
    except Exception as e:
        return html_error_fragment(f"Error loading vendors: {str(e)}")

# -------- VENDOR DETAILS (repoint) ----------
@main.route("/api/vendors/<int:vendor_id>/details")
def vendor_details(vendor_id):
    """Vendor detail using Products + usage via ConfigurationProduct."""
    try:
        vendor = Vendor.query.get_or_404(vendor_id)
        area_map = {}
        usage_rows = db.session.query(
            Product.id, Product.name,
            FunctionalArea.name.label('fa_name')
        ).join(ConfigurationProduct, ConfigurationProduct.product_id == Product.id) \
         .join(Configuration, Configuration.id == ConfigurationProduct.configuration_id) \
         .join(Function, Function.id == Configuration.function_id) \
         .join(FunctionalArea, FunctionalArea.id == Function.functional_area_id) \
         .filter(Product.vendor_id == vendor_id) \
         .distinct().all()
        used_product_ids = set()
        for pid, pname, fa_name in usage_rows:
            area_map.setdefault(fa_name, []).append((pid, pname))
            used_product_ids.add(pid)
        unused_products = Product.query.filter(Product.vendor_id == vendor_id, ~Product.id.in_(used_product_ids)).all()
        if unused_products:
            area_map.setdefault('Unassigned', []).extend([(p.id, p.name) for p in unused_products])
        versions_q = db.session.query(ProductVersion).join(Product).filter(Product.vendor_id == vendor_id)
        total_versions = versions_q.count()
        eol_soon = versions_q.filter(
            ProductVersion.support_end_date.isnot(None),
            ProductVersion.support_end_date <= func.date(func.now(), '+90 day')
        ).count()
        vendor.products_by_area = {k: [pn for _, pn in v] for k, v in area_map.items()}
        vendor.total_products = Product.query.filter_by(vendor_id=vendor_id).count()
        vendor.active_usage = db.session.query(func.count(func.distinct(Configuration.id))) \
            .join(ConfigurationProduct, ConfigurationProduct.configuration_id == Configuration.id) \
            .join(Product, Product.id == ConfigurationProduct.product_id) \
            .filter(Product.vendor_id == vendor_id).scalar()
        vendor.total_versions = total_versions
        vendor.eol_soon_versions = eol_soon
        return render_template('fragments/vendor_details.html', vendor=vendor)
    except Exception as e:
        return html_error_fragment(f"Error loading vendor details: {str(e)}")

# -------- VENDOR FORMS ----------
@main.route("/api/vendors/form")
def vendor_form():
    try:
        form = VendorForm()
        return render_template('fragments/vendor_form.html', form=form, vendor=None)
    except Exception as e:
        return html_error_fragment(f"Error loading form: {str(e)}")

@main.route("/api/vendors/<int:vendor_id>/form")
def vendor_edit_form(vendor_id):
    try:
        vendor = Vendor.query.get_or_404(vendor_id)
        form = VendorForm()
        form.populate_from_vendor(vendor)
        return render_template('fragments/vendor_form.html', form=form, vendor=vendor)
    except Exception as e:
        return html_error_fragment(f"Error loading edit form: {str(e)}")

# -------- VENDOR DELETE (guard on product & usage) ----------
@main.route("/api/vendors/<int:vendor_id>", methods=['DELETE'])
@admin_required
def delete_vendor(vendor_id):
    try:
        vendor = Vendor.query.get_or_404(vendor_id)
        name = vendor.name
        usage_exists = db.session.query(ConfigurationProduct.id) \
            .join(Product, Product.id == ConfigurationProduct.product_id) \
            .filter(Product.vendor_id == vendor_id).first()
        if usage_exists:
            return json_error_response(f"Cannot delete vendor '{name}' – products in active configurations.")
        product_count = Product.query.filter_by(vendor_id=vendor_id).count()
        if product_count > 0:
            return json_error_response(
                f"Cannot delete vendor '{name}' because it still has {product_count} products. Delete or reassign products first."
            )
        db.session.delete(vendor)
        db.session.commit()
        return api_ok({"message": f"Vendor '{name}' deleted successfully"})
    except Exception as e:
        db.session.rollback()
        return json_error_response(f"Error deleting vendor: {str(e)}")

# -------- VENDOR STATS (repoint) ----------
@main.route("/api/vendors/stats")
def vendors_stats():
    """Aggregate vendor statistics using products + configuration usage."""
    try:
        total_vendors = Vendor.query.count()
        total_products = Product.query.count()
        vendors_with_usage = db.session.query(func.count(func.distinct(Vendor.id))) \
            .join(Product, Product.vendor_id == Vendor.id) \
            .join(ConfigurationProduct, ConfigurationProduct.product_id == Product.id) \
            .join(Configuration, Configuration.id == ConfigurationProduct.configuration_id).scalar()
        avg_products_per_vendor = round(total_products / total_vendors, 1) if total_vendors else 0
        most_used = db.session.query(
            Vendor.name,
            func.count(func.distinct(Configuration.id)).label('cfg_use')
        ).join(Product, Product.vendor_id == Vendor.id) \
         .join(ConfigurationProduct, ConfigurationProduct.product_id == Product.id) \
         .join(Configuration, Configuration.id == ConfigurationProduct.configuration_id) \
         .group_by(Vendor.id, Vendor.name) \
         .order_by(func.count(func.distinct(Configuration.id)).desc()) \
         .first()
        stats = {
            'total_vendors': total_vendors,
            'vendors_with_usage': vendors_with_usage or 0,
            'avg_products_per_vendor': avg_products_per_vendor,
            'most_used_vendor': most_used.name if most_used else 'N/A',
            'most_used_vendor_cfgs': most_used.cfg_use if most_used else 0
        }
        return api_ok(stats)
    except Exception as e:
        return json_error_response(f"Error getting vendor stats: {str(e)}")

# -------- VENDOR PERFORMANCE (new metrics) ----------
@main.route("/api/vendors/performance")
def vendor_performance():
    """Vendor performance metrics based on product versions & usage."""
    try:
        most_versions = db.session.query(
            Vendor.name,
            func.count(ProductVersion.id).label('ver_count')
        ).join(Product, Product.vendor_id == Vendor.id) \
         .join(ProductVersion, ProductVersion.product_id == Product.id) \
         .group_by(Vendor.id, Vendor.name) \
         .order_by(func.count(ProductVersion.id).desc()) \
         .first()
        recent_release = db.session.query(
            Vendor.name,
            func.max(ProductVersion.release_date).label('latest_release')
        ).join(Product, Product.vendor_id == Vendor.id) \
         .join(ProductVersion, ProductVersion.product_id == Product.id) \
         .group_by(Vendor.id, Vendor.name) \
         .order_by(func.max(ProductVersion.release_date).desc().nullslast()) \
         .first()
        most_used = db.session.query(
            Vendor.name,
            func.count(func.distinct(Configuration.id)).label('cfg_use')
        ).join(Product, Product.vendor_id == Vendor.id) \
         .join(ConfigurationProduct, ConfigurationProduct.product_id == Product.id) \
         .join(Configuration, Configuration.id == ConfigurationProduct.configuration_id) \
         .group_by(Vendor.id, Vendor.name) \
         .order_by(func.count(func.distinct(Configuration.id)).desc()) \
         .first()
        return api_ok({
            'most_versions': most_versions.name if most_versions else 'N/A',
            'most_versions_count': most_versions.ver_count if most_versions else 0,
            'most_recent_release_vendor': recent_release.name if recent_release else 'N/A',
            'most_recent_release_date': recent_release.latest_release.isoformat() if (recent_release and recent_release.latest_release) else None,
            'most_used_vendor': most_used.name if most_used else 'N/A',
            'most_used_vendor_cfgs': most_used.cfg_use if most_used else 0
        })
    except Exception as e:
        return json_error_response(f"Error getting vendor performance: {str(e)}")

# -------- AGENCIES LIST (new endpoint) ----------
@main.route("/api/agencies/list")
def agencies_list_fragment():
    """HTMX fragment: agencies list with synthetic implementation counts from Configurations."""
    try:
        search = (request.args.get('search') or '').strip()
        # Subquery for configuration counts per agency
        cfg_sub = db.session.query(
            Configuration.agency_id.label('a_id'),
            func.count(Configuration.id).label('cfg_count')
        ).group_by(Configuration.agency_id).subquery()
        q = db.session.query(Agency, func.coalesce(cfg_sub.c.cfg_count, 0).label('cfg_count')) \
            .outerjoin(cfg_sub, cfg_sub.c.a_id == Agency.id)
        if search:
            q = q.filter(Agency.name.ilike(f"%{search}%"))
        agencies = q.order_by(Agency.name.asc()).all()
        # Attach synthetic attribute used by legacy template (function_implementations|length)
        result = []
        for agency, cfg_count in agencies:
            agency.function_implementations = [None] * cfg_count  # length drives display
            result.append(agency)
        return render_template('fragments/agency_list.html', agencies=result)
    except Exception as e:
        return html_error_fragment(f"Error loading agencies: {str(e)}")

@main.route("/api/agencies/stats")
def agencies_stats():
    """Aggregate agency stats using Configuration model (keeps legacy key names)."""
    try:
        total_agencies = Agency.query.count()
        active_cfgs = Configuration.query.count()
        avg_impl = round(active_cfgs / total_agencies, 1) if total_agencies else 0
        # Average vendors per agency (vendors whose products appear in that agency's configurations)
        vendor_counts_rows = db.session.query(
            Configuration.agency_id,
            func.count(func.distinct(Vendor.id)).label('v_count')
        ).join(ConfigurationProduct, ConfigurationProduct.configuration_id == Configuration.id) \
         .join(Product, Product.id == ConfigurationProduct.product_id) \
         .join(Vendor, Vendor.id == Product.vendor_id) \
         .group_by(Configuration.agency_id).all()
        avg_vendors = round(sum(r.v_count for r in vendor_counts_rows) / len(vendor_counts_rows), 1) if vendor_counts_rows else 0
        return api_ok({
            'active_implementations': active_cfgs,
            'avg_implementations_per_agency': avg_impl,
            'avg_vendors_per_agency': avg_vendors
        })
    except Exception as e:
        return json_error_response(f"Error getting agency stats: {str(e)}")

@main.route("/api/agencies/insights")
def agencies_insights():
    """Insights: tech leader (most configurations), common functional area, top vendor."""
    try:
        tech_leader = db.session.query(
            Agency.name, func.count(Configuration.id).label('cfg_count')
        ).join(Configuration, Configuration.agency_id == Agency.id) \
         .group_by(Agency.id).order_by(func.count(Configuration.id).desc()).first()
        common_area = db.session.query(
            FunctionalArea.name, func.count(Configuration.id).label('cfg_count')
        ).join(Function, Function.functional_area_id == FunctionalArea.id) \
         .join(Configuration, Configuration.function_id == Function.id) \
         .group_by(FunctionalArea.id).order_by(func.count(Configuration.id).desc()).first()
        top_vendor = db.session.query(
            Vendor.name, func.count(func.distinct(Configuration.id)).label('cfg_use')
        ).join(Product, Product.vendor_id == Vendor.id) \
         .join(ConfigurationProduct, ConfigurationProduct.product_id == Product.id) \
         .join(Configuration, Configuration.id == ConfigurationProduct.configuration_id) \
         .group_by(Vendor.id).order_by(func.count(func.distinct(Configuration.id)).desc()).first()
        return api_ok({
            'tech_leader': tech_leader.name if tech_leader else 'N/A',
            'common_area': common_area.name if common_area else 'N/A',
            'top_vendor': top_vendor.name if top_vendor else 'N/A'
        })
    except Exception as e:
        return json_error_response(f"Error getting agency insights: {str(e)}")

@main.route("/api/agencies/<int:agency_id>/details")
def agency_details_fragment(agency_id):
    """HTMX fragment: agency details panel (adapts to legacy template expectations)."""
    try:
        agency = Agency.query.get_or_404(agency_id)
        cfg_count = db.session.query(func.count(Configuration.id)).filter(Configuration.agency_id == agency_id).scalar()
        agency.function_implementations = [None] * cfg_count
        return render_template('fragments/agency_details.html', agency=agency)
    except Exception as e:
        return html_error_fragment(f"Error loading agency details: {str(e)}")

@main.route("/api/agencies/form")
def agency_form_fragment():
    try:
        form = AgencyForm()
        return render_template('fragments/agency_form.html', form=form, agency=None)
    except Exception as e:
        return html_error_fragment(f"Error loading agency form: {str(e)}")

@main.route("/api/agencies/<int:agency_id>/form")
def agency_edit_form_fragment(agency_id):
    try:
        agency = Agency.query.get_or_404(agency_id)
        form = AgencyForm()
        form.populate_from_agency(agency)
        return render_template('fragments/agency_form.html', form=form, agency=agency)
    except Exception as e:
        return html_error_fragment(f"Error loading agency edit form: {str(e)}")

@main.route('/agencies/<int:agency_id>/update', methods=['POST'])
@admin_required
def update_agency(agency_id):
    """Handle agency edit form submission (non-AJAX)."""
    agency = Agency.query.get_or_404(agency_id)
    form = AgencyForm()
    if form.validate_on_submit():
        try:
            form.populate_agency(agency)
            agency.short_name = request.form.get('short_name') or None
            db.session.commit()
            return redirect(url_for('agency.index'))
        except Exception as e:
            db.session.rollback()
            return html_error_fragment(f"Error updating agency: {str(e)}")
    # If validation fails re-render form fragment
    return render_template('fragments/agency_form.html', form=form, agency=agency)

@main.route("/api/vendors", methods=['POST'])
@admin_required
def create_vendor():
    """Create a new vendor (JSON response for HTMX form)."""
    form = VendorForm()
    if form.validate_on_submit():
        try:
            vendor = Vendor()
            form.populate_vendor(vendor)
            db.session.add(vendor)
            db.session.commit()
            return api_ok({"id": vendor.id, "message": f"Vendor '{vendor.name}' created"})
        except IntegrityError as ie:
            db.session.rollback()
            return json_validation_error_response("Duplicate vendor", {"name": "A vendor with this name already exists."})
        except Exception as e:
            db.session.rollback()
            return json_error_response(f"Error creating vendor: {str(e)}")
    return json_form_error_response(form)


@main.route("/api/vendors/<int:vendor_id>", methods=['POST'])
@admin_required
def update_vendor(vendor_id):
    """Update an existing vendor (JSON response for HTMX form)."""
    vendor = Vendor.query.get_or_404(vendor_id)
    form = VendorForm()
    if form.validate_on_submit():
        try:
            form.populate_vendor(vendor)
            db.session.commit()
            return api_ok({"id": vendor.id, "message": f"Vendor '{vendor.name}' updated"})
        except IntegrityError:
            db.session.rollback()
            return json_validation_error_response("Duplicate vendor", {"name": "A vendor with this name already exists."})
        except Exception as e:
            db.session.rollback()
            return json_error_response(f"Error updating vendor: {str(e)}")
    return json_form_error_response(form)
